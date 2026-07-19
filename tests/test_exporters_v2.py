"""Tests: V2 exports open correctly and match the compute() result."""

from __future__ import annotations

import base64
import io
import json
import unittest

from openpyxl import load_workbook

from planner_app.api import compute, default_payload
from planner_app.exporters_v2 import (
    ANNUAL_DETAIL_SPEC,
    MONEY_FMT,
    MONEY_RED_FMT,
    build_advisor_workbook,
    export_advisor_xlsx_b64,
)


def _col(header: str) -> int:
    """1-based column index of a header in the annual detail sheets."""
    for idx, (label, _getter, _fmt) in enumerate(ANNUAL_DETAIL_SPEC, start=1):
        if label == header:
            return idx
    raise KeyError(header)


def _find_row(ws, label: str, column: int = 1) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=column).value == label:
            return row
    raise KeyError(label)


class AdvisorWorkbookTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.payload = default_payload()
        cls.result = compute(cls.payload)
        cls.meta = {"generatedAt": "2026-07-19T12:00:00Z", "realDollars": True}
        cls.data = build_advisor_workbook(cls.result, cls.payload, cls.meta)
        cls.wb = load_workbook(io.BytesIO(cls.data))

    def test_sheet_order(self) -> None:
        names = [entry["scenarioName"] for entry in self.result["scenarios"]]
        self.assertEqual(
            self.wb.sheetnames,
            ["Cover", "Comparison", "Current Position"] + names + ["Charts", "Assumptions", "Sources"],
        )

    def test_charts_sheet_has_native_charts(self) -> None:
        # openpyxl may drop charts on re-read; assert on the saved archive itself.
        import zipfile

        with zipfile.ZipFile(io.BytesIO(self.data)) as zf:
            charts = [name for name in zf.namelist() if name.startswith("xl/charts/chart")]
        self.assertGreaterEqual(len(charts), 2)

    def test_cover_has_hash_date_and_assumptions(self) -> None:
        ws = self.wb["Cover"]
        row = _find_row(ws, "Input hash")
        self.assertEqual(ws.cell(row=row, column=2).value, self.result["inputHash"])
        row = _find_row(ws, "Prepared")
        self.assertEqual(ws.cell(row=row, column=2).value, "2026-07-19")
        row = _find_row(ws, "Portfolio annual return")
        self.assertAlmostEqual(ws.cell(row=row, column=2).value, 0.07)
        self.assertEqual(ws.cell(row=row, column=2).number_format, "0.0%")
        # Every scenario appears with a route summary.
        for entry in self.result["scenarios"]:
            row = _find_row(ws, entry["scenarioName"])
            self.assertTrue(ws.cell(row=row, column=2).value)

    def test_comparison_metrics_and_formats(self) -> None:
        ws = self.wb["Comparison"]
        self.assertEqual(ws.cell(row=1, column=2).value, "Path A")
        self.assertEqual(ws.freeze_panes, "B2")
        row = _find_row(ws, "Final portfolio (nominal)")
        cell = ws.cell(row=row, column=2)
        self.assertAlmostEqual(cell.value, self.result["scenarios"][0]["metrics"]["finalPortfolio"])
        self.assertEqual(cell.number_format, MONEY_FMT)

    def test_comparison_deltas_match(self) -> None:
        ws = self.wb["Comparison"]
        comparisons = self.result["comparison"]["comparisons"]
        self.assertTrue(comparisons)
        row = _find_row(ws, "Final portfolio Δ (nominal)")
        cell = ws.cell(row=row, column=2)
        self.assertAlmostEqual(cell.value, comparisons[0]["finalPortfolioDelta"])
        self.assertEqual(cell.number_format, MONEY_RED_FMT)
        # Drivers block carries each named driver.
        first_driver = next(iter(comparisons[0]["drivers"]))
        row = _find_row(ws, f"  {first_driver}")
        self.assertAlmostEqual(
            ws.cell(row=row, column=2).value, comparisons[0]["drivers"][first_driver]
        )

    def test_comparison_milestones_match(self) -> None:
        ws = self.wb["Comparison"]
        milestones = self.result["scenarios"][0]["metrics"]["milestones"]
        self.assertTrue(milestones)
        header = _find_row(ws, "Portfolio at milestones")
        row = _find_row(ws, milestones[0]["label"])
        self.assertGreater(row, header)
        self.assertAlmostEqual(ws.cell(row=row, column=2).value, milestones[0]["portfolio"])

    def test_annual_sheet_full_breakdown(self) -> None:
        entry = self.result["scenarios"][0]
        ws = self.wb[entry["scenarioName"]]
        self.assertEqual(ws.freeze_panes, "D2")
        self.assertEqual(ws.max_row, 1 + len(entry["projection"]))
        self.assertEqual(ws.max_column, len(ANNUAL_DETAIL_SPEC))
        first = entry["projection"][0]
        self.assertEqual(ws.cell(row=2, column=_col("Year")).value, first["calendarYear"])
        self.assertAlmostEqual(
            ws.cell(row=2, column=_col("Military pension")).value,
            first["incomeBreakdown"]["pension"],
        )
        self.assertAlmostEqual(
            ws.cell(row=2, column=_col("Federal tax")).value,
            first["taxBreakdown"]["federalTax"],
        )
        self.assertAlmostEqual(
            ws.cell(row=2, column=_col("Brokerage")).value,
            first["accountBalances"]["brokerage"],
        )
        last = entry["projection"][-1]
        last_row = 1 + len(entry["projection"])
        self.assertAlmostEqual(
            ws.cell(row=last_row, column=_col("Portfolio (real 2026$)")).value,
            last["portfolio"] * last["realDollarFactor"],
        )
        cf_cell = ws.cell(row=2, column=_col("Net cash flow"))
        self.assertEqual(cf_cell.number_format, MONEY_RED_FMT)

    def test_b64_wrapper_roundtrips(self) -> None:
        arg = json.dumps({"result": self.result, "payload": self.payload, "meta": self.meta})
        b64 = export_advisor_xlsx_b64(arg)
        wb = load_workbook(io.BytesIO(base64.b64decode(b64)))
        self.assertIn("Cover", wb.sheetnames)
        self.assertIn("Comparison", wb.sheetnames)

    def test_current_position_net_worth_matches_seed(self) -> None:
        from planner_app.manual_finance import flatten_manual_finance_group

        ws = self.wb["Current Position"]

        def group_total(group: str) -> float:
            flat = flatten_manual_finance_group(self.payload["manualInputs"][group])
            return sum(
                float(item.get("amountMonthly", item.get("amount")) or 0) for item in flat
            )

        row = _find_row(ws, "Total assets")
        self.assertAlmostEqual(ws.cell(row=row, column=2).value, group_total("assets"))
        row = _find_row(ws, "Net worth")
        self.assertAlmostEqual(
            ws.cell(row=row, column=2).value, group_total("assets") - group_total("debts")
        )
        row = _find_row(ws, "Net monthly cash flow")
        self.assertAlmostEqual(
            ws.cell(row=row, column=2).value, group_total("income") - group_total("expenses")
        )

    def test_assumptions_lists_overrides_with_original(self) -> None:
        label = "Employer 401(k) match (effective rate of salary)"
        payload = dict(self.payload)
        payload["referenceOverrides"] = [
            {
                "domain": "v2_benefit_rules",
                "id": "employer_match_effective_default",
                "field": "valuePercent",
                "value": 0.06,
            }
        ]
        result = compute(payload)
        wb = load_workbook(io.BytesIO(build_advisor_workbook(result, payload, self.meta)))
        ws = wb["Assumptions"]
        row = _find_row(ws, "v2_benefit_rules")
        self.assertEqual(ws.cell(row=row, column=2).value, label)
        self.assertAlmostEqual(ws.cell(row=row, column=4).value, 0.04)  # original
        self.assertAlmostEqual(ws.cell(row=row, column=5).value, 0.06)  # current
        # And the sources sheet marks the record as overridden.
        sources = wb["Sources"]
        for row_idx in range(4, sources.max_row + 1):
            if sources.cell(row=row_idx, column=2).value == label:
                self.assertIn("overridden", sources.cell(row=row_idx, column=5).value)
                break
        else:
            self.fail(f"{label!r} not found in Sources sheet")

    def test_no_overrides_shows_default_note(self) -> None:
        ws = self.wb["Assumptions"]
        row = _find_row(ws, "Your overrides")
        self.assertIn("sourced default", ws.cell(row=row + 1, column=1).value)

    def test_sources_sheet_has_cited_records_with_links(self) -> None:
        ws = self.wb["Sources"]
        self.assertGreater(ws.max_row, 10)
        linked = 0
        for row_idx in range(4, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=4)
            if cell.value:
                linked += 1
                self.assertTrue(str(cell.value).startswith("http"), cell.value)
                self.assertIsNotNone(cell.hyperlink)
        self.assertGreater(linked, 0)


if __name__ == "__main__":
    unittest.main()
