"""V2 export: the advisor-grade XLSX workbook.

Consumes a `compute()` result (planner_app.api) so the export always matches
the dashboard, and embeds the deterministic input hash to detect drift. Used
natively (tests, CLI) and in the browser via Pyodide (openpyxl installed via
micropip on first use). Also consumes the compute *payload* (manual inputs,
reference overrides, scenarios) so it can document the client's current
position and every assumption behind the projection.

The HTML report export is client-side (web/src/report/) — it renders the
actual app components, so it is not duplicated here.
"""

from __future__ import annotations

import base64
import io
import json
from typing import Any, Callable

MONEY_FMT = '"$"#,##0'
MONEY_RED_FMT = '"$"#,##0;[Red]-"$"#,##0'
PCT_FMT = "0.0%"
DISCLAIMER = "Planning estimate with simplified effective-rate taxes — not financial advice."

METRIC_ROWS = [
    ("finalPortfolio", "Final portfolio (nominal)"),
    ("finalPortfolioReal", "Final portfolio (real 2026$)"),
    ("totalGrossIncome", "Lifetime gross income"),
    ("totalTaxFreeIncome", "Lifetime tax-free income"),
    ("totalTaxes", "Lifetime taxes"),
    ("totalHealthcareCost", "Lifetime healthcare"),
    ("totalEmployerMatch", "Lifetime employer match"),
    ("lifetimePensionValue", "Lifetime pension paid"),
    ("lifetimeSocialSecurity", "Lifetime Social Security"),
    ("totalUnfundedSpending", "Unfunded spending"),
    ("pensionStartYear", "Pension starts"),
    ("ssStartYear", "Social Security starts"),
    ("withdrawalEligibleYear", "Retirement withdrawals eligible"),
    ("depletionAge", "Portfolio depletion age"),
]

# Metrics that are years/ages, not dollars.
_YEAR_METRICS = {"pensionStartYear", "ssStartYear", "withdrawalEligibleYear", "depletionAge"}


# ---------- advisor workbook ----------

def _sub(outer: str, key: str) -> Callable[[dict[str, Any]], Any]:
    return lambda row: (row.get(outer) or {}).get(key, 0)


def _portfolio_withdrawals(row: dict[str, Any]) -> float:
    w = row.get("withdrawals") or {}
    return sum(v for k, v in w.items() if k != "taxesOnWithdrawals" and isinstance(v, (int, float)))


def _real_portfolio(row: dict[str, Any]) -> float:
    return (row.get("portfolio") or 0) * (row.get("realDollarFactor") or 1)


# (header, getter, number format or None) — the full annual cash-flow statement.
ANNUAL_DETAIL_SPEC: list[tuple[str, Callable[[dict[str, Any]], Any], str | None]] = [
    ("Year", lambda r: r.get("calendarYear"), None),
    ("Age", lambda r: r.get("age"), None),
    ("Phase", lambda r: r.get("phaseLabel"), None),
    ("Military base pay", _sub("incomeBreakdown", "militaryBasePay"), MONEY_FMT),
    ("BAH (housing)", _sub("incomeBreakdown", "militaryBah"), MONEY_FMT),
    ("BAS (subsistence)", _sub("incomeBreakdown", "militaryBas"), MONEY_FMT),
    ("VA compensation", _sub("incomeBreakdown", "vaCompensation"), MONEY_FMT),
    ("Military pension", _sub("incomeBreakdown", "pension"), MONEY_FMT),
    ("Salary", _sub("incomeBreakdown", "salaryBase"), MONEY_FMT),
    ("PhD stipend", _sub("incomeBreakdown", "phdStipend"), MONEY_FMT),
    ("GI Bill housing", _sub("incomeBreakdown", "giBillHousing"), MONEY_FMT),
    ("GI Bill books", _sub("incomeBreakdown", "giBillBooks"), MONEY_FMT),
    ("Social Security", _sub("incomeBreakdown", "socialSecurity"), MONEY_FMT),
    ("Gross income", lambda r: r.get("grossIncome"), MONEY_FMT),
    ("Tax-free income", lambda r: r.get("taxFreeIncome"), MONEY_FMT),
    ("Total income", lambda r: r.get("totalIncome"), MONEY_FMT),
    ("Federal tax", _sub("taxBreakdown", "federalTax"), MONEY_FMT),
    ("State tax", _sub("taxBreakdown", "stateTax"), MONEY_FMT),
    ("Withdrawal tax", _sub("taxBreakdown", "withdrawalTax"), MONEY_FMT),
    ("Total taxes", lambda r: r.get("taxes"), MONEY_FMT),
    ("Healthcare", lambda r: r.get("healthcareCost"), MONEY_FMT),
    ("Living expenses", lambda r: r.get("livingExpenses"), MONEY_FMT),
    ("Retirement contributions", lambda r: r.get("retirementSavings"), MONEY_FMT),
    ("Employer match", lambda r: r.get("employerMatch"), MONEY_FMT),
    ("Surplus invested", lambda r: r.get("positiveSurplusInvested"), MONEY_FMT),
    ("Net cash flow", lambda r: r.get("netCashFlow"), MONEY_RED_FMT),
    ("Unfunded spending", lambda r: r.get("unfundedSpending"), MONEY_RED_FMT),
    ("Portfolio withdrawals", _portfolio_withdrawals, MONEY_FMT),
    ("RMD", lambda r: r.get("rmd"), MONEY_FMT),
    ("Cash", _sub("accountBalances", "cash"), MONEY_FMT),
    ("Brokerage", _sub("accountBalances", "brokerage"), MONEY_FMT),
    ("Roth IRA", _sub("accountBalances", "rothIra"), MONEY_FMT),
    ("TSP Roth", _sub("accountBalances", "tspRoth"), MONEY_FMT),
    ("Traditional 401(k)", _sub("accountBalances", "trad401k"), MONEY_FMT),
    ("Total portfolio", lambda r: r.get("portfolio"), MONEY_FMT),
    ("Portfolio (real 2026$)", _real_portfolio, MONEY_FMT),
]


def _set(cell, value, *, fmt: str | None = None, bold: bool = False, size: int | None = None):
    cell.value = value
    if fmt:
        cell.number_format = fmt
    if bold or size:
        from openpyxl.styles import Font

        cell.font = Font(bold=bold, size=size or 11)
    return cell


def _header_row(ws, row: int, labels: list[str], start_col: int = 1):
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill(start_color="FFEEF1F5", end_color="FFEEF1F5", fill_type="solid")
    for offset, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + offset, value=label)
        cell.font = Font(bold=True)
        cell.fill = fill


_BLOCK_TYPE_LABELS = {
    "grad_school": "Grad school",
    "research_career": "Research career",
    "tech_career": "Tech career",
    "gap_year": "Gap year",
}


def _route_summary(scenario: dict[str, Any]) -> str:
    parts: list[str] = []
    exit_info = scenario.get("serviceExit") or {}
    if exit_info:
        exit_label = str(exit_info.get("type", "separation")).replace("_", " ")
        parts.append(f"Service to {exit_info.get('year', '?')} ({exit_label})")
    for block in scenario.get("blocks") or []:
        block_type = str(block.get("type", "block"))
        label = _BLOCK_TYPE_LABELS.get(block_type, block_type.replace("_", " ").capitalize())
        months = block.get("durationMonths")
        if months:
            label += f" ({months / 12:.0f}y)" if months % 12 == 0 else f" ({months} mo)"
        parts.append(label)
    return " → ".join(parts)


def _find(records: list[dict[str, Any]], record_id: str) -> dict[str, Any]:
    for record in records:
        if record.get("id") == record_id:
            return record
    return {}


def _policy_rows(payload: dict[str, Any]) -> list[tuple[str, Any, str | None]]:
    """Key model assumptions, after applying the user's reference overrides."""
    from .api import apply_reference_overrides

    domains, _tables = apply_reference_overrides(payload.get("referenceOverrides") or [])
    growth = _find(domains.get("investment_policies", []), "portfolio_growth_core")
    rows: list[tuple[str, Any, str | None]] = [
        ("Portfolio annual return", growth.get("annualReturnRate"), PCT_FMT),
        ("Surplus investment rate", growth.get("surplusInvestmentRate"), PCT_FMT),
        ("Retirement withdrawal rate", growth.get("withdrawalRate"), PCT_FMT),
    ]
    for rule_id, label in [
        ("inflation_general_default", "General inflation"),
        ("employer_match_effective_default", "Employer 401(k) match (effective)"),
        ("capital_gains_rate_default", "Capital gains rate"),
        ("ss_cola_default", "Social Security COLA"),
        ("military_raise_default", "Military annual raise"),
    ]:
        record = _find(domains.get("v2_benefit_rules", []), rule_id)
        if record:
            rows.append((label, record.get("valuePercent"), PCT_FMT))
    for rule_id, label in [
        ("va_cola", "VA compensation COLA"),
        ("living_expense_growth_default", "Living expense growth"),
    ]:
        record = _find(domains.get("benefit_rules", []), rule_id)
        if record:
            rows.append((label, record.get("valuePercent"), PCT_FMT))
    return rows


def _build_cover_sheet(ws, result: dict[str, Any], payload: dict[str, Any], meta: dict[str, Any]):
    ws.title = "Cover"
    _set(ws.cell(row=1, column=1), "Career Plan Codex", bold=True, size=18)
    _set(ws.cell(row=2, column=1), "Financial plan workbook — career path comparison")

    generated = (meta or {}).get("generatedAt") or ""
    _set(ws.cell(row=4, column=1), "Prepared", bold=True)
    ws.cell(row=4, column=2, value=str(generated)[:10])
    _set(ws.cell(row=5, column=1), "Input hash", bold=True)
    ws.cell(row=5, column=2, value=result.get("inputHash", ""))
    _set(ws.cell(row=6, column=1), "Values", bold=True)
    ws.cell(row=6, column=2, value="Nominal dollars unless labelled real (2026$)")

    row = 8
    _set(ws.cell(row=row, column=1), "Paths compared", bold=True, size=13)
    row += 1
    _header_row(ws, row, ["Path", "Route", "Notes"])
    for entry in result["scenarios"]:
        row += 1
        ws.cell(row=row, column=1, value=entry["scenarioName"])
        ws.cell(row=row, column=2, value=_route_summary(entry.get("scenario") or {}))
        ws.cell(row=row, column=3, value=(entry.get("scenario") or {}).get("notes") or "")

    row += 2
    _set(ws.cell(row=row, column=1), "Key assumptions", bold=True, size=13)
    for label, value, fmt in _policy_rows(payload):
        row += 1
        ws.cell(row=row, column=1, value=label)
        _set(ws.cell(row=row, column=2), value, fmt=fmt)

    row += 2
    ws.cell(row=row, column=1, value=DISCLAIMER)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 60


def _build_comparison_sheet(ws, result: dict[str, Any]):
    from openpyxl.utils import get_column_letter

    ws.title = "Comparison"
    scenarios = result["scenarios"]
    names = [entry["scenarioName"] for entry in scenarios]
    _header_row(ws, 1, ["Metric"] + names)
    row = 1
    for key, label in METRIC_ROWS:
        row += 1
        ws.cell(row=row, column=1, value=label)
        for col, entry in enumerate(scenarios, start=2):
            fmt = None if key in _YEAR_METRICS else MONEY_FMT
            _set(ws.cell(row=row, column=col), entry["metrics"].get(key), fmt=fmt)

    comparisons = (result.get("comparison") or {}).get("comparisons") or []
    by_id = {entry["scenarioId"]: entry["scenarioName"] for entry in scenarios}
    if comparisons:
        baseline_name = by_id.get(comparisons[0].get("baselineScenarioId"), "baseline")
        row += 2
        _set(ws.cell(row=row, column=1), f"Versus baseline — {baseline_name}", bold=True, size=13)
        row += 1
        _header_row(ws, row, [""] + [by_id.get(c["scenarioId"], c["scenarioId"]) for c in comparisons])
        delta_rows: list[tuple[str, Callable[[dict[str, Any]], Any], str | None]] = [
            ("Final portfolio Δ (nominal)", lambda c: c.get("finalPortfolioDelta"), MONEY_RED_FMT),
            ("Final portfolio Δ (real 2026$)", lambda c: c.get("finalPortfolioRealDelta"), MONEY_RED_FMT),
            ("Lifetime taxes Δ", lambda c: c.get("totalTaxesDelta"), MONEY_RED_FMT),
            ("Lifetime healthcare Δ", lambda c: c.get("healthcareCostDelta"), MONEY_RED_FMT),
            ("Breakeven year", lambda c: c.get("breakevenYear") or "never", None),
            (
                "Biggest driver",
                lambda c: (c.get("biggestDriver") or {}).get("label", "—"),
                None,
            ),
        ]
        for label, getter, fmt in delta_rows:
            row += 1
            ws.cell(row=row, column=1, value=label)
            for col, comp in enumerate(comparisons, start=2):
                _set(ws.cell(row=row, column=col), getter(comp), fmt=fmt)

        driver_labels: list[str] = []
        for comp in comparisons:
            for label in (comp.get("drivers") or {}):
                if label not in driver_labels:
                    driver_labels.append(label)
        if driver_labels:
            row += 2
            _set(ws.cell(row=row, column=1), "Cumulative drivers vs baseline", bold=True)
            for label in driver_labels:
                row += 1
                ws.cell(row=row, column=1, value=f"  {label}")
                for col, comp in enumerate(comparisons, start=2):
                    _set(ws.cell(row=row, column=col), (comp.get("drivers") or {}).get(label), fmt=MONEY_RED_FMT)

    milestone_labels: list[str] = []
    for entry in scenarios:
        for milestone in entry["metrics"].get("milestones") or []:
            if milestone["label"] not in milestone_labels:
                milestone_labels.append(milestone["label"])
    if milestone_labels:
        for title, field in [
            ("Portfolio at milestones", "portfolio"),
            ("Sustainable 4% withdrawal at milestones", "withdrawalAt4Pct"),
        ]:
            row += 2
            _set(ws.cell(row=row, column=1), title, bold=True, size=13)
            row += 1
            _header_row(ws, row, [""] + names)
            for label in milestone_labels:
                row += 1
                ws.cell(row=row, column=1, value=label)
                for col, entry in enumerate(scenarios, start=2):
                    match = next(
                        (m for m in entry["metrics"].get("milestones") or [] if m["label"] == label),
                        None,
                    )
                    _set(ws.cell(row=row, column=col), match.get(field) if match else None, fmt=MONEY_FMT)

    row += 2
    ws.cell(row=row, column=1, value=f"Input hash: {result.get('inputHash', '')}")

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 34
    for col in range(2, len(scenarios) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _build_annual_sheet(wb, entry: dict[str, Any]):
    from openpyxl.utils import get_column_letter

    title = (entry["scenarioName"] or entry["scenarioId"])[:31]
    ws = wb.create_sheet(title=title)
    _header_row(ws, 1, [header for header, _getter, _fmt in ANNUAL_DETAIL_SPEC])
    for row_idx, row in enumerate(entry["projection"], start=2):
        for col, (_header, getter, fmt) in enumerate(ANNUAL_DETAIL_SPEC, start=1):
            _set(ws.cell(row=row_idx, column=col), getter(row), fmt=fmt)
    ws.freeze_panes = "D2"
    widths = {1: 8, 2: 6, 3: 24}
    for col in range(1, len(ANNUAL_DETAIL_SPEC) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 15)
    return ws


def _amount(item: dict[str, Any]) -> float:
    value = item.get("amountMonthly", item.get("amount"))
    return float(value or 0)


def _shown(item: dict[str, Any]) -> bool:
    """Mirror the Finances screen: hide zero-amount show_only_if_used rows."""
    return bool(_amount(item)) or item.get("displayMode") != "show_only_if_used"


def _position_block(ws, row: int, sections: list[dict[str, Any]], total_label: str) -> tuple[int, float]:
    """Write one statement block (items grouped by section) and return its total."""
    from .manual_finance import flatten_manual_finance_group

    flat = flatten_manual_finance_group(sections or [])
    total = 0.0
    current_section = None
    for item in flat:
        total += _amount(item)
        if not _shown(item):
            continue
        if item.get("sectionLabel") != current_section:
            current_section = item.get("sectionLabel")
            _set(ws.cell(row=row, column=1), current_section, bold=True)
            row += 1
        ws.cell(row=row, column=1, value=f"  {item.get('label', item.get('id'))}")
        _set(ws.cell(row=row, column=2), _amount(item), fmt=MONEY_FMT)
        ws.cell(row=row, column=3, value=item.get("notes") or "")
        row += 1
    _set(ws.cell(row=row, column=1), total_label, bold=True)
    _set(ws.cell(row=row, column=2), total, fmt=MONEY_RED_FMT, bold=True)
    return row + 1, total


def _build_position_sheet(ws, payload: dict[str, Any]):
    """Net-worth and monthly cash-flow statements from the manual baseline."""
    from .seed_data import MANUAL_CASHFLOW_SEED

    ws.title = "Current Position"
    manual = payload.get("manualInputs") or MANUAL_CASHFLOW_SEED

    _set(ws.cell(row=1, column=1), "Current financial position", bold=True, size=14)
    row = 3
    _set(ws.cell(row=row, column=1), "Net worth statement", bold=True, size=13)
    row += 1
    _header_row(ws, row, ["Item", "Amount", "Notes"])
    row, assets = _position_block(ws, row + 1, manual.get("assets") or [], "Total assets")
    row += 1
    row, debts = _position_block(ws, row, manual.get("debts") or [], "Total debts")
    row += 1
    _set(ws.cell(row=row, column=1), "Net worth", bold=True, size=13)
    _set(ws.cell(row=row, column=2), assets - debts, fmt=MONEY_RED_FMT, bold=True)

    row += 3
    _set(ws.cell(row=row, column=1), "Monthly cash flow", bold=True, size=13)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            "Manual baseline only — reference-backed income (military pay, salary) is "
            "projected by the engine; see the per-path annual sheets."
        ),
    )
    row += 1
    _header_row(ws, row, ["Item", "Monthly", "Notes"])
    row, income = _position_block(ws, row + 1, manual.get("income") or [], "Total monthly income")
    row += 1
    row, expenses = _position_block(ws, row, manual.get("expenses") or [], "Total monthly expenses")
    row += 1
    _set(ws.cell(row=row, column=1), "Net monthly cash flow", bold=True, size=13)
    _set(ws.cell(row=row, column=2), income - expenses, fmt=MONEY_RED_FMT, bold=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50


def _build_assumptions_sheet(ws, result: dict[str, Any], payload: dict[str, Any]):
    ws.title = "Assumptions"
    _set(ws.cell(row=1, column=1), "Assumptions & overrides", bold=True, size=14)

    row = 3
    _set(ws.cell(row=row, column=1), "Model assumptions", bold=True, size=13)
    for label, value, fmt in _policy_rows(payload):
        row += 1
        ws.cell(row=row, column=1, value=label)
        _set(ws.cell(row=row, column=2), value, fmt=fmt)

    scenarios = result.get("scenarios") or []
    if scenarios:
        row += 2
        _set(ws.cell(row=row, column=1), "Per-path settings", bold=True, size=13)
        row += 1
        _header_row(ws, row, [""] + [entry["scenarioName"] for entry in scenarios])
        settings: list[tuple[str, Callable[[dict[str, Any]], Any]]] = [
            ("Service exit", lambda s: _route_summary({"serviceExit": s.get("serviceExit")}) or "—"),
            ("VA disability used", lambda s: "Yes" if s.get("useVa") else "No"),
            ("VA rating", lambda s: s.get("selectedVaRatingId") or "—"),
            ("GI Bill used", lambda s: "Yes" if s.get("useGiBill") else "No"),
            ("SS claim age", lambda s: (s.get("retirement") or {}).get("ssClaimAge")),
            ("Withdrawal age", lambda s: (s.get("retirement") or {}).get("withdrawalAgeYears")),
            ("Withdrawal policy", lambda s: (s.get("retirement") or {}).get("withdrawalPolicy")),
        ]
        for label, getter in settings:
            row += 1
            ws.cell(row=row, column=1, value=label)
            for col, entry in enumerate(scenarios, start=2):
                ws.cell(row=row, column=col, value=getter(entry.get("scenario") or {}))

    from .api import apply_reference_overrides

    overrides = payload.get("referenceOverrides") or []
    row += 2
    _set(ws.cell(row=row, column=1), "Your overrides", bold=True, size=13)
    row += 1
    if not overrides:
        ws.cell(row=row, column=1, value="None — every reference value is at its sourced default.")
    else:
        pristine_domains, pristine_tables = apply_reference_overrides([])
        _header_row(ws, row, ["Domain", "Record", "Field", "Original", "Current"])
        for override in overrides:
            row += 1
            domain = override.get("domain", "")
            record = _find(
                pristine_domains.get(domain, []) or pristine_tables.get(domain, []),
                override.get("id", ""),
            )
            ws.cell(row=row, column=1, value=domain)
            ws.cell(row=row, column=2, value=record.get("label") or override.get("id"))
            ws.cell(row=row, column=3, value=override.get("field"))
            ws.cell(row=row, column=4, value=record.get(override.get("field", ""), "—"))
            ws.cell(row=row, column=5, value=override.get("value"))

    ws.column_dimensions["A"].width = 34
    for letter in ("B", "C", "D", "E"):
        ws.column_dimensions[letter].width = 22


def _build_sources_sheet(ws, payload: dict[str, Any]):
    """Every cited reference record — the same rows the Sources screen shows."""
    from .api import bootstrap_data

    ws.title = "Sources"
    overridden = {
        (o.get("domain"), o.get("id")) for o in payload.get("referenceOverrides") or []
    }
    _set(ws.cell(row=1, column=1), "Reference sources", bold=True, size=14)
    _header_row(ws, 3, ["Domain", "Record", "Source", "URL", "Status"])
    row = 3
    domains = bootstrap_data()["referenceDomains"]
    for domain in sorted(domains):
        for record in domains[domain]:
            if not (record.get("sourceLabel") or record.get("sourceUrl")):
                continue
            row += 1
            ws.cell(row=row, column=1, value=domain)
            ws.cell(row=row, column=2, value=record.get("label") or record.get("id"))
            ws.cell(row=row, column=3, value=record.get("sourceLabel") or "")
            url = record.get("sourceUrl") or ""
            cell = ws.cell(row=row, column=4, value=url)
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"
            status = record.get("verificationStatus") or ""
            if (domain, record.get("id")) in overridden:
                status = (status + " (overridden)").strip()
            ws.cell(row=row, column=5, value=status)
    ws.freeze_panes = "A4"
    widths = {"A": 26, "B": 34, "C": 44, "D": 56, "E": 22}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def build_advisor_workbook(
    result: dict[str, Any],
    payload: dict[str, Any] | None = None,
    meta: dict[str, Any] | None = None,
) -> bytes:
    """Advisor-grade workbook: Cover, Comparison, Current Position, full annual
    detail per path, Assumptions & overrides, and Sources."""
    from openpyxl import Workbook

    payload = payload or {}
    meta = meta or {}

    wb = Workbook()
    _build_cover_sheet(wb.active, result, payload, meta)
    _build_comparison_sheet(wb.create_sheet(), result)
    _build_position_sheet(wb.create_sheet(), payload)
    for entry in result["scenarios"]:
        _build_annual_sheet(wb, entry)
    _build_assumptions_sheet(wb.create_sheet(), result, payload)
    _build_sources_sheet(wb.create_sheet(), payload)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_advisor_xlsx_b64(arg_json: str) -> str:
    """Pyodide-friendly wrapper: JSON `{result, payload, meta}` in, base64 XLSX out."""
    arg = json.loads(arg_json)
    result = arg.get("result") or arg  # tolerate a bare compute() result
    return base64.b64encode(
        build_advisor_workbook(result, arg.get("payload"), arg.get("meta"))
    ).decode("ascii")

