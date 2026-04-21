from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CLAIMS_SHEET = "claims"
LOOKUP_SHEET = "lookup"
INSTRUCTIONS_SHEET = "instructions"
SOURCE_SLOT_COUNT = 3
CHANGE_TYPE_VALUES = {
    "new_fill",
    "value_correction",
    "source_correction",
    "status_correction",
    "staleness_update",
    "case_split",
    "keep_current",
}
REVIEW_STATUS_VALUES = {
    "pending",
    "needs_second_review",
    "corroborated",
    "accepted",
    "rejected",
    "superseded",
}
DECISION_ACTION_VALUES = {
    "apply_to_claim",
    "keep_current",
    "create_case_specific_followup",
    "request_manual_review",
}
RESOLUTION_STATUS_VALUES = {
    "official_verified",
    "public_verified",
    "estimated",
    "not_publicly_available",
    "deferred",
}
EVIDENCE_TIER_VALUES = {
    "official",
    "public",
    "secondary",
    "estimate",
    "not_publicly_available",
}
CONFIDENCE_VALUES = {"high", "medium", "low"}
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
HEADER_FONT = Font(bold=True)

CLAIM_EXPORT_COLUMNS = [
    "claim_id",
    "section_id",
    "section",
    "domain",
    "record_id",
    "target_label",
    "field",
    "field_label",
    "value_kind",
    "baseline_value",
    "current_value",
    "proposed_value",
    "current_status",
    "current_verification_status",
    "research_note",
    "resolution_status",
    "evidence_tier",
    "confidence",
    "estimate_rationale",
]
for slot in range(1, SOURCE_SLOT_COUNT + 1):
    CLAIM_EXPORT_COLUMNS.extend(
        [
            f"source_{slot}_title",
            f"source_{slot}_url",
            f"source_{slot}_publisher",
            f"source_{slot}_published_date",
            f"source_{slot}_accessed_date",
            f"source_{slot}_excerpt",
        ]
    )

LOOKUP_EXPORT_COLUMNS = [
    "section_id",
    "domain",
    "record_id",
    "label",
    "verification_status",
    "field_count",
    "record_json",
]

CONFLICT_EXPORT_COLUMNS = [
    "conflict_id",
    "claim_id",
    "submission_id",
    "submitted_by",
    "submitted_role",
    "submitted_model",
    "submitted_at",
    "accepted_state_hash",
    "accepted_value_snapshot",
    "accepted_resolution_status",
    "accepted_evidence_tier",
    "accepted_source_fingerprint",
    "proposed_value",
    "proposed_resolution_status",
    "proposed_evidence_tier",
    "proposed_confidence",
    "proposed_estimate_rationale",
    "proposed_research_note",
]
for slot in range(1, SOURCE_SLOT_COUNT + 1):
    CONFLICT_EXPORT_COLUMNS.extend(
        [
            f"proposed_source_{slot}_title",
            f"proposed_source_{slot}_url",
            f"proposed_source_{slot}_publisher",
            f"proposed_source_{slot}_published_date",
            f"proposed_source_{slot}_accessed_date",
            f"proposed_source_{slot}_excerpt",
        ]
    )
CONFLICT_EXPORT_COLUMNS.extend(
    [
        "change_type",
        "disagreement_summary",
        "corroborates_submission_id",
        "review_status",
        "decision_action",
        "decision_by",
        "decision_at",
        "decision_notes",
    ]
)

REVIEW_QUEUE_COLUMNS = CONFLICT_EXPORT_COLUMNS + ["queue_reason"]

INSTRUCTIONS_ROWS = [
    ("Purpose", "Use the claims sheet as the working research database for Reference Data."),
    ("Editing", "Edit proposed_value, resolution_status, evidence_tier, confidence, estimate_rationale, and source_* columns."),
    ("Conflict Log", "Do not overwrite a populated accepted row when you materially disagree. Append the disagreement to reference_research_conflicts.csv instead."),
    ("File Roles", "Accepted-state edits belong in reference_research_claims.csv. Disagreements and adjudication belong in reference_research_conflicts.csv. The resolved and review-queue CSVs are derived/read-only."),
    ("Resolution Status", "Use one of: official_verified, public_verified, estimated, not_publicly_available, deferred."),
    ("Evidence Tier", "Use one of: official, public, secondary, estimate, not_publicly_available."),
    ("Confidence", "Use high, medium, or low for estimated rows."),
    ("Percent Inputs", "Percent values accept 0.028, 2.8, or 2.8%."),
    ("Boolean Inputs", "Boolean flags accept true/false, yes/no, enabled/disabled, or with/without dependents."),
    ("Reset Behavior", "Blank proposed_value removes the imported research override for that claim on the next import."),
    ("Import Scope", "Import replaces prior research-import values, claim statuses, and manual research citations."),
    ("Multi-Agent Review", "Use the accepted claims CSV for non-conflicting fills. Use the conflict log for disagreements, corrections, corroboration, and reviewer decisions."),
    ("Derived Files", "Do not edit reference_research_claims_resolved.csv or reference_research_review_queue.csv directly. Regenerate them with python3 app.py reconcile-reference-research."),
    ("Final Upload", "After final adjudication and reconcile, upload reference_research_claims_resolved.csv back into the app."),
]

REFERENCE_RESEARCH_PROMPT = """You are helping fill a structured Reference Data research file for a financial planning app.

Your job is to research each row carefully and populate only the research-facing fields while preserving the file schema exactly.

Your priority is trustworthiness, not speed or coverage. Do not fill a value unless you can support it with strong evidence. A blank field with a correct status is better than a weak or speculative answer.

Core objective:
- For each row, determine the best publicly supportable value for the claim.
- Prefer official and primary sources over convenience sources.
- If a value cannot be directly verified, do not guess. Either provide a clearly labeled estimate with rationale or mark it as not publicly available / deferred.

Schema rules:
- Do not rename columns.
- Do not add or delete columns.
- Do not change row order.
- Do not alter these identifier/context columns:
  - `claim_id`
  - `section_id`
  - `domain`
  - `record_id`
  - `field`
  - `field_label`
  - `value_kind`
  - `baseline_value`
  - `current_value`
- Only fill or update these research-facing columns:
  - `proposed_value`
  - `research_note`
  - `resolution_status`
  - `evidence_tier`
  - `confidence`
  - `estimate_rationale`
  - `source_1_title`
  - `source_1_url`
  - `source_1_publisher`
  - `source_1_published_date`
  - `source_1_accessed_date`
  - `source_1_excerpt`
  - `source_2_*`
  - `source_3_*`

Source quality rules:
- Use the highest-authority source reasonably available.
- Source preference order:
  1. Official government, military, VA, IRS, DoD, .gov, statute, regulation, or official public program source
  2. Official institution source such as a university, employer, insurer, or program administrator
  3. Official published documents such as benefit tables, program handbooks, tuition pages, fee schedules, or policy manuals
  4. Reputable public datasets or recognized nonpartisan research organizations
  5. Secondary sources only when primary sources are unavailable, and only if they are credible and specific
- Do not use low-trust sources as the basis for a final claim unless the task is explicitly an estimate and no better source exists.
- Avoid or heavily discount:
  - AI-generated summaries
  - blogs and marketing pages
  - SEO listicles
  - unsourced comparison sites
  - forums and social media
  - user-edited pages unless they link directly to a primary source
- If a secondary source cites a primary source, prefer the primary source directly.

Research discipline:
- Verify that the source actually supports the exact row-level claim, not just the general topic.
- Prefer the most recent authoritative source when claims are time-sensitive.
- If multiple credible sources conflict, choose the most authoritative and current one and briefly note the conflict in `research_note`.
- Do not silently average conflicting figures.
- Do not infer missing values unless the estimate rules below are followed.
- Do not use stale values when a newer official value is available.
- If a row appears jurisdiction-specific, year-specific, or conditional, capture the most relevant current value and note important scope conditions in `research_note`.

Allowed resolution statuses:
- `official_verified`
- `public_verified`
- `estimated`
- `not_publicly_available`
- `deferred`

Allowed evidence tiers:
- `official`
- `public`
- `secondary`
- `estimate`
- `not_publicly_available`

How to choose status:
- Use `official_verified` when the claim is supported by an official primary source.
- Use `public_verified` when supported by a strong public but non-official source or institution source.
- Use `estimated` only when no direct public value exists but a reasonable estimate can be derived from public inputs.
- Use `not_publicly_available` when the value is not meaningfully available from public sources.
- Use `deferred` when the row needs more investigation or the evidence is too ambiguous to resolve safely.

Estimate rules:
- Only estimate when a direct value is unavailable.
- Estimates must be conservative and based on public inputs.
- For every estimate, you must fill:
  - `resolution_status` = `estimated`
  - `evidence_tier` = `estimate`
  - `confidence` = `high`, `medium`, or `low`
  - `estimate_rationale` with a short explanation of the derivation
- Do not disguise an estimate as verified data.
- If the estimate is weak or highly assumption-dependent, prefer `deferred` or `not_publicly_available`.

Citation rules:
- Use up to 3 sources when helpful.
- Prefer 1 strong primary source over 3 weak sources.
- Keep `source_*_excerpt` short, specific, and tied to the exact claim.
- Use direct source URLs whenever possible, not search-result URLs.
- Fill `source_*_published_date` when available.
- Fill `source_*_accessed_date` for every cited source if possible.
- Keep publisher names specific and recognizable.

Formatting rules:
- Preserve the file structure exactly.
- Keep dates in `YYYY-MM-DD` when available.
- Keep values in the unit/format implied by the row and existing schema.
- For percentage fields, do not switch between whole-percent and decimal forms unless clearly required by the existing row format.
- For text fields, use concise factual wording.
- Do not insert commentary outside the designated columns.

Decision rules:
- Accuracy is more important than completeness.
- If unsure, do not force a value.
- If the source does not clearly support the claim, leave `proposed_value` blank and set the appropriate status.
- If a better source is likely to exist but you cannot confidently locate it, use `deferred` rather than a weak citation.
- Do not overwrite a current value just because a different number appears online; change it only when supported by better evidence.

Working method:
- Read the row context carefully:
  - `section_id`
  - `domain`
  - `record_id`
  - `field`
  - `field_label`
  - `value_kind`
  - `research_note`
- Determine what exact fact is being claimed.
- Find the best available source.
- Confirm the value, scope, timing, and units.
- Populate the research-facing columns only.
- Preserve all unchanged rows.

Output requirement:
- Return the completed file in the exact same tabular structure with the same rows and columns.
- Do not summarize instead of editing.
- Do not drop rows.
- Do not reorder columns unless the file format forces it and all original columns are preserved.

Multi-agent disagreement rules:
- Treat `reference_research_claims.csv` as the current accepted-state file.
- If a claim row is blank, unresolved, or you are only strengthening the existing accepted sources without changing the accepted value/status, you may edit `reference_research_claims.csv` directly.
- If you materially disagree with a nonblank accepted row’s value, status, estimate rationale, or sources, do not overwrite that row directly.
- Instead, append a new row to `reference_research_conflicts.csv`.
- Fill `claim_id` exactly as it appears in the accepted claims CSV.
- Use `change_type` to classify the disagreement: `new_fill`, `value_correction`, `source_correction`, `status_correction`, `staleness_update`, `case_split`, or `keep_current`.
- If you are independently confirming a prior correction, append a new conflict row and set `corroborates_submission_id` to the earlier submission.
- Do not average conflicting values unless the field is explicitly an aggregate metric and the average is methodologically justified.
- If two values appear valid for different contexts, submit a `case_split` conflict instead of forcing one canonical answer.
- Leave reviewer decision fields alone unless you are acting as the adjudicating reviewer.
- The app should import only the resolved claims CSV after reconciliation, not raw unresolved conflicts.

Four-file workflow rules:
- You may receive 4 CSV files:
  - `reference_research_claims.csv`
  - `reference_research_conflicts.csv`
  - `reference_research_claims_resolved.csv`
  - `reference_research_review_queue.csv`
- Treat the files as having different roles:
  - `reference_research_claims.csv` is the accepted-state working file.
  - `reference_research_conflicts.csv` is the append-only disagreement and review log.
  - `reference_research_claims_resolved.csv` is a derived output produced after reconciliation.
  - `reference_research_review_queue.csv` is a derived queue of unresolved or stale conflicts produced after reconciliation.
- Do not edit `reference_research_claims_resolved.csv` directly.
- Do not edit `reference_research_review_queue.csv` directly.
- Use `reference_research_review_queue.csv` only to identify which conflict rows still need corroboration or reviewer action.
- Make your edits only in:
  - `reference_research_claims.csv`
  - `reference_research_conflicts.csv`

How to use `reference_research_claims.csv`:
- Edit this file directly only when:
  - the claim is blank or unresolved, or
  - you are strengthening the accepted row without materially changing the accepted value, status, or claim meaning
- If you materially disagree with a populated accepted row, do not overwrite it directly.
- A material disagreement includes:
  - a different value
  - a different resolution status
  - a different estimate rationale
  - replacing the accepted source basis with a competing source basis

How to use `reference_research_conflicts.csv`:
- Append one new row for each disagreement, correction, corroboration, staleness update, or case split.
- Do not delete prior conflict rows.
- If you agree with an earlier correction after independent review, append a new row and set `corroborates_submission_id` to the earlier submission.
- Use:
  - `change_type = new_fill` when proposing a first populated answer through the conflict workflow
  - `value_correction` for a different value
  - `source_correction` for a different citation basis
  - `status_correction` for a different status
  - `staleness_update` when accepted data appears outdated
  - `case_split` when multiple values appear valid under different conditions
  - `keep_current` when explicitly supporting the current accepted state against a competing proposal
- Fill these provenance fields when possible:
  - `submission_id`
  - `submitted_by`
  - `submitted_role`
  - `submitted_model`
  - `submitted_at`
- Copy `claim_id` exactly from the accepted claims file.
- Copy the accepted row context when possible:
  - `accepted_value_snapshot`
  - `accepted_resolution_status`
  - `accepted_evidence_tier`
- If you cannot reliably compute `accepted_state_hash` or `accepted_source_fingerprint`, leave them blank rather than inventing them.

Review and adjudication rules:
- If you are acting as a researcher, leave reviewer decision fields unchanged unless you are explicitly asked to adjudicate.
- If you are acting as a reviewer/adjudicator, update these fields in `reference_research_conflicts.csv`:
  - `review_status`
  - `decision_action`
  - `decision_by`
  - `decision_at`
  - `decision_notes`
- Allowed `review_status` values:
  - `pending`
  - `needs_second_review`
  - `corroborated`
  - `accepted`
  - `rejected`
  - `superseded`
- Allowed `decision_action` values:
  - `apply_to_claim`
  - `keep_current`
  - `create_case_specific_followup`
  - `request_manual_review`
- Do not directly modify the accepted claims row to resolve a disagreement; record the decision in the conflict log and let reconciliation produce the resolved CSV.

Case-split and averaging rules:
- Do not average conflicting values by default.
- Only average if the field is explicitly an aggregate metric and averaging is methodologically correct.
- If multiple values are valid for different jurisdictions, years, applicant types, or conditions, use `case_split` instead of forcing one canonical answer.
- If a case split is needed, explain the split clearly in `disagreement_summary` and recommend `decision_action = create_case_specific_followup`.

Output requirement for multi-pass review:
- Return updated versions of the files you actually changed.
- Usually this means:
  - updated `reference_research_claims.csv`
  - updated `reference_research_conflicts.csv`
- Do not rewrite derived files unless explicitly asked to regenerate them outside the spreadsheet workflow.
- The final file to upload back into the app is `reference_research_claims_resolved.csv` after reconciliation.
"""


def stringify_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, separators=(",", ":"))
    return str(value)


def normalize_token(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace("%", " percent")
        .replace("/", " ")
        .replace("-", "_")
        .replace(" ", "_")
    )


def normalize_resolution_status(value: Any) -> str:
    token = normalize_token(value)
    if token in {"", "none"}:
        return ""
    aliases = {
        "official": "official_verified",
        "official_verified": "official_verified",
        "public": "public_verified",
        "public_verified": "public_verified",
        "estimated": "estimated",
        "estimate": "estimated",
        "not_publicly_available": "not_publicly_available",
        "not_public": "not_publicly_available",
        "deferred": "deferred",
    }
    normalized = aliases.get(token, token)
    if normalized not in RESOLUTION_STATUS_VALUES:
        raise ValueError(f"Unsupported resolution_status: {value}")
    return normalized


def normalize_evidence_tier(value: Any, resolution_status: str = "") -> str:
    token = normalize_token(value)
    if token in {"", "none"}:
        if resolution_status == "official_verified":
            return "official"
        if resolution_status == "public_verified":
            return "public"
        if resolution_status == "estimated":
            return "estimate"
        if resolution_status == "not_publicly_available":
            return "not_publicly_available"
        return ""
    aliases = {
        "official": "official",
        "public": "public",
        "secondary": "secondary",
        "estimate": "estimate",
        "estimated": "estimate",
        "not_publicly_available": "not_publicly_available",
    }
    normalized = aliases.get(token, token)
    if normalized not in EVIDENCE_TIER_VALUES:
        raise ValueError(f"Unsupported evidence_tier: {value}")
    return normalized


def normalize_confidence(value: Any) -> str:
    token = normalize_token(value)
    if token in {"", "none"}:
        return ""
    aliases = {
        "high": "high",
        "medium": "medium",
        "med": "medium",
        "low": "low",
    }
    normalized = aliases.get(token, token)
    if normalized not in CONFIDENCE_VALUES:
        raise ValueError(f"Unsupported confidence: {value}")
    return normalized


def placeholder_status_for_resolution(resolution_status: str) -> str:
    if resolution_status in {"official_verified", "public_verified"}:
        return "resolved"
    if resolution_status == "estimated":
        return "estimated"
    if resolution_status == "not_publicly_available":
        return "not_publicly_available"
    if resolution_status == "deferred":
        return "deferred"
    return "source_pending"


def parse_research_value(raw_value: Any, kind: str | None) -> Any:
    if raw_value is None:
        return None
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return None
    else:
        text = str(raw_value).strip()
        if text == "":
            return None

    normalized_kind = kind or "text"
    if normalized_kind in {"currency", "number"}:
        return float(text.replace("$", "").replace(",", ""))
    if normalized_kind == "percent":
        cleaned = text.replace(",", "")
        if cleaned.endswith("%"):
            return float(cleaned[:-1]) / 100
        numeric = float(cleaned)
        if abs(numeric) > 1:
            return numeric / 100
        return numeric
    if normalized_kind in {"year", "paygrade"}:
        return int(float(text.upper().replace("E-", "")))
    if normalized_kind in {"boolean_flag", "dependents_flag"}:
        token = normalize_token(text)
        truthy = {"1", "true", "yes", "enabled", "with_dependents"}
        falsy = {"0", "false", "no", "disabled", "without_dependents"}
        if token in truthy:
            return 1
        if token in falsy:
            return 0
        raise ValueError(f"Unsupported boolean-style value: {raw_value}")
    if normalized_kind == "json":
        return json.loads(text)
    if normalized_kind == "date":
        return text
    return text


def build_reference_research_rows(bootstrap: dict[str, Any]) -> list[dict[str, Any]]:
    records_by_key = {
        (domain, record["id"]): record
        for domain, records in (bootstrap.get("referenceDomains") or {}).items()
        for record in records
    }
    research_import_values = {
        (item["domain"], item["recordId"], item["field"]): item["value"]
        for item in bootstrap.get("referenceOverrides", [])
        if item.get("scope") == "research_import"
    }
    rows: list[dict[str, Any]] = []
    for claim in bootstrap.get("referencedValues", []):
        record = records_by_key.get((claim["targetDomain"], claim["targetRecordId"]), {})
        baseline_value = record.get("baselineValues", {}).get(claim["targetField"], claim.get("currentValue"))
        proposed_value = research_import_values.get((claim["targetDomain"], claim["targetRecordId"], claim["targetField"]))
        row = {
            "claim_id": claim["id"],
            "section_id": claim.get("sectionId", ""),
            "section": claim.get("section", ""),
            "domain": claim["targetDomain"],
            "record_id": claim["targetRecordId"],
            "target_label": claim.get("friendlyTargetLabel", ""),
            "field": claim["targetField"],
            "field_label": claim.get("fieldLabel", ""),
            "value_kind": claim.get("valueKind") or "",
            "baseline_value": stringify_cell_value(baseline_value),
            "current_value": stringify_cell_value(claim.get("currentValue")),
            "proposed_value": stringify_cell_value(proposed_value),
            "current_status": claim.get("status", ""),
            "current_verification_status": claim.get("verificationStatus", ""),
            "research_note": claim.get("researchNote", ""),
            "resolution_status": claim.get("verificationStatus", "") if claim.get("verificationStatus") in RESOLUTION_STATUS_VALUES else "",
            "evidence_tier": claim.get("evidenceTier", ""),
            "confidence": claim.get("confidence", ""),
            "estimate_rationale": claim.get("estimateRationale", ""),
        }
        source_links = claim.get("sourceLinks", [])[:SOURCE_SLOT_COUNT]
        for slot in range(1, SOURCE_SLOT_COUNT + 1):
            link = source_links[slot - 1] if slot - 1 < len(source_links) else {}
            row[f"source_{slot}_title"] = link.get("title", "")
            row[f"source_{slot}_url"] = link.get("url", "")
            row[f"source_{slot}_publisher"] = link.get("publisher", "")
            row[f"source_{slot}_published_date"] = link.get("publishedDate", "")
            row[f"source_{slot}_accessed_date"] = link.get("accessedDate", "")
            row[f"source_{slot}_excerpt"] = link.get("noteExcerpt", "")
        rows.append(row)
    return rows


def build_reference_lookup_rows(bootstrap: dict[str, Any]) -> list[dict[str, Any]]:
    metadata = bootstrap.get("referenceFieldMetadata", {})
    rows: list[dict[str, Any]] = []
    for domain, records in (bootstrap.get("referenceDomains") or {}).items():
        researchable_fields = metadata.get(domain, {}).get("researchableFields", [])
        for record in records:
            rows.append(
                {
                    "section_id": metadata.get(domain, {}).get("sectionId", ""),
                    "domain": domain,
                    "record_id": record["id"],
                    "label": record.get("label") or record.get("schoolName") or record["id"],
                    "verification_status": record.get("verificationStatus", ""),
                    "field_count": len(researchable_fields),
                    "record_json": json.dumps(record, separators=(",", ":"), ensure_ascii=True),
                }
            )
    return rows


def _apply_sheet_header_style(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells[:50]]
        width = max((len(value) for value in values), default=12)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 42)


def write_reference_research_workbook(
    *,
    bootstrap: dict[str, Any],
    output_path: str | Path,
) -> tuple[Path, list[dict[str, Any]]]:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    rows = build_reference_research_rows(bootstrap)
    lookup_rows = build_reference_lookup_rows(bootstrap)

    workbook = Workbook()
    claims_sheet = workbook.active
    claims_sheet.title = CLAIMS_SHEET
    claims_sheet.append(CLAIM_EXPORT_COLUMNS)
    for row in rows:
        claims_sheet.append([stringify_cell_value(row.get(column, "")) for column in CLAIM_EXPORT_COLUMNS])
    _apply_sheet_header_style(claims_sheet)
    _autosize_columns(claims_sheet)

    lookup_sheet = workbook.create_sheet(LOOKUP_SHEET)
    lookup_sheet.append(LOOKUP_EXPORT_COLUMNS)
    for row in lookup_rows:
        lookup_sheet.append([stringify_cell_value(row.get(column, "")) for column in LOOKUP_EXPORT_COLUMNS])
    _apply_sheet_header_style(lookup_sheet)
    _autosize_columns(lookup_sheet)

    instructions_sheet = workbook.create_sheet(INSTRUCTIONS_SHEET)
    instructions_sheet.append(["topic", "details"])
    for row in INSTRUCTIONS_ROWS:
        instructions_sheet.append(list(row))
    _apply_sheet_header_style(instructions_sheet)
    _autosize_columns(instructions_sheet)

    workbook.save(output)
    return output, rows


def write_reference_research_csv(
    *,
    rows: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CLAIM_EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return output


def write_reference_research_prompt(
    *,
    output_path: str | Path,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(REFERENCE_RESEARCH_PROMPT, encoding="utf-8")
    return output


def load_reference_research_rows(path: str | Path) -> list[dict[str, Any]]:
    input_path = Path(path)
    if input_path.suffix.lower() == ".csv":
        with input_path.open("r", encoding="utf-8", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    workbook = load_workbook(input_path, data_only=True)
    if CLAIMS_SHEET not in workbook.sheetnames:
        raise ValueError(f"{input_path} does not contain a '{CLAIMS_SHEET}' sheet")
    sheet = workbook[CLAIMS_SHEET]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        payload = {headers[index]: row[index] for index in range(len(headers))}
        if str(payload.get("claim_id") or "").strip():
            rows.append(payload)
    return rows


def write_reference_research_conflict_csv(
    *,
    rows: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CONFLICT_EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return output


def write_reference_research_review_queue_csv(
    *,
    rows: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_QUEUE_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return output


def ensure_reference_research_conflict_csv(output_path: str | Path) -> Path:
    output = Path(output_path)
    if output.exists():
        return output
    return write_reference_research_conflict_csv(rows=[], output_path=output)


def load_reference_research_conflict_rows(path: str | Path) -> list[dict[str, Any]]:
    input_path = Path(path)
    if not input_path.exists():
        return []
    if input_path.suffix.lower() != ".csv":
        raise ValueError("Conflict logs must be stored as CSV files.")
    with input_path.open("r", encoding="utf-8", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle) if any(str(value or "").strip() for value in row.values())]


def normalize_change_type(value: Any) -> str:
    token = normalize_token(value)
    if token in {"", "none"}:
        return ""
    normalized = token
    if normalized not in CHANGE_TYPE_VALUES:
        raise ValueError(f"Unsupported change_type: {value}")
    return normalized


def normalize_review_status(value: Any) -> str:
    token = normalize_token(value)
    if token in {"", "none"}:
        return "pending"
    normalized = token
    if normalized not in REVIEW_STATUS_VALUES:
        raise ValueError(f"Unsupported review_status: {value}")
    return normalized


def normalize_decision_action(value: Any) -> str:
    token = normalize_token(value)
    if token in {"", "none"}:
        return ""
    normalized = token
    if normalized not in DECISION_ACTION_VALUES:
        raise ValueError(f"Unsupported decision_action: {value}")
    return normalized


def _source_columns_for_prefix(prefix: str) -> list[str]:
    columns: list[str] = []
    for slot in range(1, SOURCE_SLOT_COUNT + 1):
        columns.extend(
            [
                f"{prefix}_{slot}_title",
                f"{prefix}_{slot}_url",
                f"{prefix}_{slot}_publisher",
                f"{prefix}_{slot}_published_date",
                f"{prefix}_{slot}_accessed_date",
                f"{prefix}_{slot}_excerpt",
            ]
        )
    return columns


def accepted_source_fingerprint(row: dict[str, Any]) -> str:
    payload = [
        str(row.get(column, "") or "").strip()
        for column in _source_columns_for_prefix("source")
    ]
    return hashlib.sha1("|".join(payload).encode("utf-8")).hexdigest()[:12]


def accepted_state_hash(row: dict[str, Any]) -> str:
    payload = {
        "claim_id": str(row.get("claim_id", "") or "").strip(),
        "current_value": stringify_cell_value(row.get("current_value", "")),
        "proposed_value": stringify_cell_value(row.get("proposed_value", "")),
        "research_note": str(row.get("research_note", "") or "").strip(),
        "resolution_status": str(row.get("resolution_status", "") or "").strip(),
        "evidence_tier": str(row.get("evidence_tier", "") or "").strip(),
        "confidence": str(row.get("confidence", "") or "").strip(),
        "estimate_rationale": str(row.get("estimate_rationale", "") or "").strip(),
        "source_fingerprint": accepted_source_fingerprint(row),
    }
    return hashlib.sha1(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()[:16]


def _proposed_source_present(row: dict[str, Any]) -> bool:
    return any(str(row.get(column, "") or "").strip() for column in _source_columns_for_prefix("proposed_source"))


def _set_proposed_sources(target_row: dict[str, Any], conflict_row: dict[str, Any]) -> None:
    for column in _source_columns_for_prefix("source"):
        target_row[column] = ""
    for slot in range(1, SOURCE_SLOT_COUNT + 1):
        for field in ("title", "url", "publisher", "published_date", "accessed_date", "excerpt"):
            target_row[f"source_{slot}_{field}"] = str(conflict_row.get(f"proposed_source_{slot}_{field}", "") or "").strip()


def _review_queue_row(conflict_row: dict[str, Any], reason: str) -> dict[str, Any]:
    queued = {column: str(conflict_row.get(column, "") or "").strip() for column in CONFLICT_EXPORT_COLUMNS}
    queued["queue_reason"] = reason
    return queued


def reconcile_reference_research_rows(
    *,
    accepted_rows: list[dict[str, Any]],
    conflict_rows: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    resolved_rows = [dict(row) for row in accepted_rows]
    rows_by_claim = {str(row.get("claim_id") or "").strip(): row for row in resolved_rows}
    applied_conflicts: list[dict[str, Any]] = []
    review_queue: list[dict[str, Any]] = []

    ordered_conflicts = sorted(
        conflict_rows,
        key=lambda row: (
            str(row.get("submitted_at") or ""),
            str(row.get("conflict_id") or ""),
            str(row.get("submission_id") or ""),
        ),
    )

    for raw_conflict in ordered_conflicts:
        conflict = {column: str(raw_conflict.get(column, "") or "").strip() for column in CONFLICT_EXPORT_COLUMNS}
        if not any(conflict.values()):
            continue
        claim_id = conflict.get("claim_id", "")
        accepted_row = rows_by_claim.get(claim_id)
        if not accepted_row:
            review_queue.append(_review_queue_row(conflict, "unknown_claim_id"))
            continue

        review_status = normalize_review_status(conflict.get("review_status"))
        decision_action = normalize_decision_action(conflict.get("decision_action"))
        normalize_change_type(conflict.get("change_type"))

        if review_status == "accepted" and decision_action == "apply_to_claim":
            expected_hash = conflict.get("accepted_state_hash", "")
            current_hash = accepted_state_hash(accepted_row)
            if expected_hash and expected_hash != current_hash:
                review_queue.append(_review_queue_row(conflict, "accepted_state_hash_mismatch"))
                continue

            if conflict.get("proposed_value", "") != "":
                accepted_row["proposed_value"] = conflict["proposed_value"]
            if conflict.get("proposed_resolution_status", "") != "":
                accepted_row["resolution_status"] = conflict["proposed_resolution_status"]
            if conflict.get("proposed_evidence_tier", "") != "":
                accepted_row["evidence_tier"] = conflict["proposed_evidence_tier"]
            if conflict.get("proposed_confidence", "") != "":
                accepted_row["confidence"] = conflict["proposed_confidence"]
            if conflict.get("proposed_estimate_rationale", "") != "":
                accepted_row["estimate_rationale"] = conflict["proposed_estimate_rationale"]
            if conflict.get("proposed_research_note", "") != "":
                accepted_row["research_note"] = conflict["proposed_research_note"]
            if _proposed_source_present(conflict):
                _set_proposed_sources(accepted_row, conflict)
            applied_conflicts.append(conflict)
            continue

        if review_status in {"pending", "needs_second_review", "corroborated"}:
            review_queue.append(_review_queue_row(conflict, "awaiting_review"))
            continue

        if review_status == "accepted" and decision_action != "apply_to_claim":
            continue

    return {
        "resolved_rows": resolved_rows,
        "applied_conflicts": applied_conflicts,
        "review_queue_rows": review_queue,
    }


def _document_id_for_source(values: tuple[str, str, str, str]) -> str:
    digest = hashlib.sha1("|".join(values).encode("utf-8")).hexdigest()[:12]
    return f"manual_doc__{digest}"


def build_research_import_bundle(
    *,
    rows: list[dict[str, Any]],
    bootstrap: dict[str, Any],
) -> dict[str, list[dict[str, Any]]]:
    claims_by_id = {item["id"]: item for item in bootstrap.get("referencedValues", [])}
    value_overrides: list[dict[str, Any]] = []
    claim_overrides: list[dict[str, Any]] = []
    documents_by_id: dict[str, dict[str, Any]] = {}
    claim_documents: list[dict[str, Any]] = []

    for row in rows:
        claim_id = str(row.get("claim_id") or "").strip()
        if not claim_id:
            continue
        claim = claims_by_id.get(claim_id)
        if not claim:
            raise ValueError(f"Unknown claim_id in import file: {claim_id}")

        resolution_status = normalize_resolution_status(row.get("resolution_status"))
        evidence_tier = normalize_evidence_tier(row.get("evidence_tier"), resolution_status)
        confidence = normalize_confidence(row.get("confidence"))
        estimate_rationale = str(row.get("estimate_rationale") or "").strip()
        research_note = str(row.get("research_note") or "").strip()
        proposed_value_raw = row.get("proposed_value")
        proposed_value = parse_research_value(proposed_value_raw, claim.get("valueKind"))
        has_non_resolution_payload = any(
            (
                evidence_tier,
                confidence,
                estimate_rationale,
                proposed_value is not None,
            )
        )
        if has_non_resolution_payload and not resolution_status:
            raise ValueError(f"{claim_id} is missing resolution_status")
        if resolution_status == "estimated" and not confidence:
            raise ValueError(f"{claim_id} requires confidence for estimated rows")
        if resolution_status == "estimated" and not estimate_rationale:
            raise ValueError(f"{claim_id} requires estimate_rationale for estimated rows")

        source_link_count = 0
        for slot in range(1, SOURCE_SLOT_COUNT + 1):
            title = str(row.get(f"source_{slot}_title") or "").strip()
            url = str(row.get(f"source_{slot}_url") or "").strip()
            publisher = str(row.get(f"source_{slot}_publisher") or "").strip()
            published_date = str(row.get(f"source_{slot}_published_date") or "").strip()
            accessed_date = str(row.get(f"source_{slot}_accessed_date") or "").strip()
            excerpt = str(row.get(f"source_{slot}_excerpt") or "").strip()
            if not any((title, url, publisher, published_date, accessed_date, excerpt)):
                continue
            if not resolution_status:
                continue
            source_link_count += 1
            document_id = _document_id_for_source((title, url, publisher, published_date))
            documents_by_id.setdefault(
                document_id,
                {
                    "id": document_id,
                    "title": title or url or f"{claim.get('fieldLineItem', claim_id)} source {slot}",
                    "publisher": publisher,
                    "url": url,
                    "sourceType": "manual_research_import",
                    "publishedDate": published_date,
                    "accessedDate": accessed_date,
                    "notes": "",
                },
            )
            claim_documents.append(
                {
                    "claimId": claim_id,
                    "documentId": document_id,
                    "role": "supporting",
                    "sortOrder": slot - 1,
                    "noteExcerpt": excerpt,
                }
            )

        has_research_payload = bool(resolution_status)

        if proposed_value is not None:
            value_overrides.append(
                {
                    "domain": claim["targetDomain"],
                    "recordId": claim["targetRecordId"],
                    "field": claim["targetField"],
                    "value": proposed_value,
                    "reason": f"Research import for {claim.get('fieldLineItem', claim_id)}",
                }
            )

        if has_research_payload:
            claim_overrides.append(
                {
                    "claimId": claim_id,
                    "researchNote": research_note,
                    "verificationStatus": resolution_status,
                    "placeholderStatus": placeholder_status_for_resolution(resolution_status),
                    "evidenceTier": evidence_tier,
                    "confidence": confidence,
                    "estimateRationale": estimate_rationale,
                }
            )

    return {
        "value_overrides": value_overrides,
        "claim_overrides": claim_overrides,
        "documents": list(documents_by_id.values()),
        "claim_documents": claim_documents,
    }
